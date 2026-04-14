# bulk_import_ui/bulk_import.py
# Bulk contract import + master data (lessors, assets, services)
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import json
import time
from core.utils import *
from conf.constants import *
from core.db import *
from core.auth import get_current_user, get_user_ip

# Contracts sheet — must match template headers and row.get() keys exactly
_HDR_ADVANCE_PAYMENT_FIXED = "Advance Payment (Fixed only)"
_HDR_RS_PAYMENT_ADVANCE = "Revenue Share Payment Advance (Revenue Share only)"
_HDR_RS_ADV_MODE = "Rev. Share Advance Mode (none/chronological/periods/spread_proportional)"


def _excel_row_effectively_empty(row) -> bool:
    for v in row:
        if pd.notna(v) and str(v).strip() != "":
            return False
    return True


def _write_instructions_sheet(wb, title: str, lines: list[str]) -> None:
    ws = wb.create_sheet("Instructions", 0)
    rows = [[title], [""]] + [[ln] for ln in lines]
    for row_num, instruction in enumerate(rows, 1):
        cell = ws.cell(row=row_num, column=1)
        cell.value = instruction[0]
        if row_num == 1:
            cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 100


def _style_master_data_headers(ws, headers: list[str], column_widths: list[float]) -> None:
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width


def generate_master_lessors_template():
    """Excel template: bulk-create lessor master records (no contracts)."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    _write_instructions_sheet(
        wb,
        "BULK LESSOR IMPORT",
        [
            "1. Use the **Lessors** sheet only.",
            "2. **Lessor Name*** is required and must be unique (not already in the system).",
            "3. Other columns are optional: Description, Tax ID, Supplier Code, IBAN.",
            "4. Delete the example rows before importing.",
            "5. Upload the file from Bulk Import → Lessors → Import.",
        ],
    )
    ws = wb.create_sheet("Lessors", 1)
    headers = [
        "Lessor Name*",
        "Description",
        "Tax ID",
        "Supplier Code",
        "IBAN",
    ]
    _style_master_data_headers(ws, headers, [28, 36, 18, 18, 28])
    examples = [
        ["Example Lessor Alpha", "Optional notes", "", "SUP-001", ""],
        ["Example Lessor Beta", "", "TAX-9", "", "EG123456789"],
    ]
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row_num, row_data in enumerate(examples, 2):
        for col_num, value in enumerate(row_data, 1):
            c = ws.cell(row=row_num, column=col_num)
            c.value = value
            c.border = thin
            if col_num == 1:
                c.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    return wb


def generate_master_assets_template():
    """Excel template: bulk-create asset master records."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    _write_instructions_sheet(
        wb,
        "BULK ASSET IMPORT",
        [
            "1. Use the **Assets** sheet only.",
            "2. **Asset Name*** is required and must be unique (not already in the system).",
            "3. **Cost Center** is optional.",
            "4. Delete the example rows before importing.",
            "5. Upload from Bulk Import → Assets → Import.",
        ],
    )
    ws = wb.create_sheet("Assets", 1)
    headers = ["Asset Name*", "Cost Center"]
    _style_master_data_headers(ws, headers, [32, 24])
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row_num, row_data in enumerate([["Example Asset North", "CC-100"]], 2):
        for col_num, value in enumerate(row_data, 1):
            c = ws.cell(row=row_num, column=col_num)
            c.value = value
            c.border = thin
            if col_num == 1:
                c.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    return wb


def generate_master_services_template():
    """Excel template: bulk-create service master records."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    _write_instructions_sheet(
        wb,
        "BULK SERVICE IMPORT",
        [
            "1. Use the **Services** sheet only.",
            "2. **Service Name*** is required and must be unique (not already in the system).",
            "3. **Currency*** must be EGP or USD.",
            "4. Description is optional.",
            "5. Delete the example rows before importing.",
            "6. Upload from Bulk Import → Services → Import.",
        ],
    )
    ws = wb.create_sheet("Services", 1)
    headers = ["Service Name*", "Description", "Currency* (EGP/USD)"]
    _style_master_data_headers(ws, headers, [28, 40, 22])
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row_num, row_data in enumerate(
        [["Maintenance — Example", "Optional description", "EGP"]], 2
    ):
        for col_num, value in enumerate(row_data, 1):
            c = ws.cell(row=row_num, column=col_num)
            c.value = value
            c.border = thin
            if col_num in (1, 3):
                c.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    return wb


def process_bulk_import_master_lessors(uploaded_file):
    """Validate and insert lessors from the Lessors sheet."""
    try:
        load_all()
        lessors_df = st.session_state.lessors_df.copy()
        df = pd.read_excel(uploaded_file, sheet_name="Lessors")
        all_errors = []
        pending = []
        seen_names = set()

        for idx, row in df.iterrows():
            row_num = idx + 2
            if _excel_row_effectively_empty(row):
                continue
            raw = row.get("Lessor Name*", "")
            name = str(raw).strip() if not pd.isna(raw) else ""
            if not name:
                all_errors.append(f"Lessors Row {row_num}: Lessor Name is required")
                continue
            key = name
            if key in seen_names:
                all_errors.append(
                    f"Lessors Row {row_num}: duplicate Lessor Name '{name}' in file"
                )
                continue
            seen_names.add(key)
            if not lessors_df.empty and not lessors_df[
                lessors_df["name"].str.strip() == name
            ].empty:
                all_errors.append(
                    f"Lessors Row {row_num}: Lessor '{name}' already exists in the system"
                )
                continue

            def _s(col):
                v = row.get(col, "")
                return "" if pd.isna(v) else str(v).strip()

            pending.append(
                {
                    "name": name,
                    "description": _s("Description"),
                    "tax_id": _s("Tax ID"),
                    "supplier_code": _s("Supplier Code"),
                    "iban": _s("IBAN"),
                }
            )

        if all_errors:
            return False, all_errors, [], 0

        working = lessors_df.copy()
        success_count = 0
        failed = []
        for item in pending:
            nid = next_int_id(working, 1)
            lessor_data = {
                "id": str(nid),
                "name": item["name"],
                "description": item["description"],
                "tax_id": item["tax_id"],
                "supplier_code": item["supplier_code"],
                "iban": item["iban"],
            }
            if insert_lessor(lessor_data):
                success_count += 1
                current_user = get_current_user()
                log_action(
                    user_id=current_user["id"] if current_user else None,
                    user_name=current_user["name"] if current_user else "System",
                    action_type="create",
                    entity_type="lessor",
                    entity_id=str(nid),
                    entity_name=item["name"],
                    action_details=f"Bulk imported lessor: {item['name']}",
                    ip_address=get_user_ip(),
                )
                working = pd.concat(
                    [working, pd.DataFrame([{"id": str(nid), "name": item["name"]}])],
                    ignore_index=True,
                )
            else:
                failed.append(f"{item['name']}: database insert failed")

        load_all()
        return True, failed, [], success_count
    except Exception as e:
        return False, [f"Error processing lessors file: {str(e)}"], [], 0


def process_bulk_import_master_assets(uploaded_file):
    """Validate and insert assets from the Assets sheet."""
    try:
        load_all()
        assets_df = st.session_state.assets_df.copy()
        df = pd.read_excel(uploaded_file, sheet_name="Assets")
        all_errors = []
        pending = []
        seen_names = set()

        for idx, row in df.iterrows():
            row_num = idx + 2
            if _excel_row_effectively_empty(row):
                continue
            raw = row.get("Asset Name*", "")
            name = str(raw).strip() if not pd.isna(raw) else ""
            if not name:
                all_errors.append(f"Assets Row {row_num}: Asset Name is required")
                continue
            if name in seen_names:
                all_errors.append(
                    f"Assets Row {row_num}: duplicate Asset Name '{name}' in file"
                )
                continue
            seen_names.add(name)
            if not assets_df.empty and not assets_df[
                assets_df["name"].str.strip() == name
            ].empty:
                all_errors.append(
                    f"Assets Row {row_num}: Asset '{name}' already exists in the system"
                )
                continue
            raw_cc = row.get("Cost Center", "")
            cost_center = "" if pd.isna(raw_cc) else str(raw_cc).strip()
            pending.append({"name": name, "cost_center": cost_center})

        if all_errors:
            return False, all_errors, [], 0

        working = assets_df.copy()
        success_count = 0
        failed = []
        for item in pending:
            nid = next_int_id(working, 101)
            asset_data = {
                "id": str(nid),
                "name": item["name"],
                "cost_center": item["cost_center"],
            }
            if insert_asset(asset_data):
                success_count += 1
                current_user = get_current_user()
                log_action(
                    user_id=current_user["id"] if current_user else None,
                    user_name=current_user["name"] if current_user else "System",
                    action_type="create",
                    entity_type="asset",
                    entity_id=str(nid),
                    entity_name=item["name"],
                    action_details=f"Bulk imported asset: {item['name']}",
                    ip_address=get_user_ip(),
                )
                working = pd.concat(
                    [working, pd.DataFrame([{"id": str(nid), "name": item["name"]}])],
                    ignore_index=True,
                )
            else:
                failed.append(f"{item['name']}: database insert failed")

        load_all()
        return True, failed, [], success_count
    except Exception as e:
        return False, [f"Error processing assets file: {str(e)}"], [], 0


def process_bulk_import_master_services(uploaded_file):
    """Validate and insert services from the Services sheet."""
    try:
        load_all()
        services_df = st.session_state.services_df.copy()
        df = pd.read_excel(uploaded_file, sheet_name="Services")
        all_errors = []
        pending = []
        seen_names = set()

        for idx, row in df.iterrows():
            row_num = idx + 2
            if _excel_row_effectively_empty(row):
                continue
            raw = row.get("Service Name*", "")
            name = str(raw).strip() if not pd.isna(raw) else ""
            if not name:
                all_errors.append(f"Services Row {row_num}: Service Name is required")
                continue
            if name in seen_names:
                all_errors.append(
                    f"Services Row {row_num}: duplicate Service Name '{name}' in file"
                )
                continue
            seen_names.add(name)
            if not services_df.empty and not services_df[
                services_df["name"].str.strip() == name
            ].empty:
                all_errors.append(
                    f"Services Row {row_num}: Service '{name}' already exists in the system"
                )
                continue
            raw_cur = row.get("Currency* (EGP/USD)", "")
            currency = str(raw_cur).strip().upper() if not pd.isna(raw_cur) else ""
            if currency not in ("EGP", "USD"):
                all_errors.append(
                    f"Services Row {row_num}: Currency must be EGP or USD"
                )
                continue
            raw_d = row.get("Description", "")
            desc = "" if pd.isna(raw_d) else str(raw_d).strip()
            pending.append({"name": name, "description": desc, "currency": currency})

        if all_errors:
            return False, all_errors, [], 0

        working = services_df.copy()
        success_count = 0
        failed = []
        for item in pending:
            nid = next_int_id(working, 1)
            service_data = {
                "id": str(nid),
                "name": item["name"],
                "description": item["description"],
                "currency": item["currency"],
            }
            if insert_service(service_data):
                success_count += 1
                current_user = get_current_user()
                log_action(
                    user_id=current_user["id"] if current_user else None,
                    user_name=current_user["name"] if current_user else "System",
                    action_type="create",
                    entity_type="service",
                    entity_id=str(nid),
                    entity_name=item["name"],
                    action_details=f"Bulk imported service: {item['name']}",
                    ip_address=get_user_ip(),
                )
                working = pd.concat(
                    [working, pd.DataFrame([{"id": str(nid), "name": item["name"]}])],
                    ignore_index=True,
                )
            else:
                failed.append(f"{item['name']}: database insert failed")

        load_all()
        return True, failed, [], success_count
    except Exception as e:
        return False, [f"Error processing services file: {str(e)}"], [], 0

def generate_bulk_import_template():
    """Generate Excel template for bulk contract import"""
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create Contracts sheet
    ws_contracts = wb.create_sheet("Contracts", 0)
    
    # Define headers
    headers = [
        "Contract Name*",
        "Contract Type* (Fixed/Revenue Share/ROU)",
        "Currency* (EGP/USD)",
        "Asset Category* (Store/Other)",
        "Asset/Store Name*",
        "Commencement Date* (YYYY-MM-DD)",
        "Tenure Years*",
        "Tenure Months* (0-11)",
        "First Payment Date (YYYY-MM-DD)",
        "Discount Rate (%) (ROU only)",
        "Tax (%)",
        "Payment Frequency* (Yearly/2 Months/Monthly/Quarter)",
        "Is Tax Added (ROU: Yes/No)",
        "Yearly Increase Type (Increased %/Fixed Amount Increased)",
        "Yearly Increase (%)",
        "Yearly Increase Fixed Amount (per period, same currency as contract)",
        "Rent Amount (EGP/month) (Fixed/ROU)",
        "Revenue Min (EGP) (Revenue Share)",
        "Revenue Max (EGP) (Revenue Share)",
        "Revenue Share % (Revenue Share)",
        "Share % After Max (Revenue Share)",
        "Sales Type (Revenue Share: Net/Total without discount)",
        _HDR_ADVANCE_PAYMENT_FIXED,
        _HDR_RS_PAYMENT_ADVANCE,
        _HDR_RS_ADV_MODE,
        "Free Months (comma-separated, Fixed/Revenue Share/ROU: e.g., 1,2,3)",
        "Advance Months (ROU / Revenue Share “periods” advance mode, e.g. 1,2,3)"
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws_contracts.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Set column widths
    column_widths = [25, 20, 15, 18, 25, 20, 15, 15, 20, 15, 10, 30, 22, 30, 15, 28, 20, 15, 15, 18, 18, 25, 22, 30, 38, 35, 40]
    for col_num, width in enumerate(column_widths, 1):
        ws_contracts.column_dimensions[get_column_letter(col_num)].width = width
    
    # Add example rows
    example_data = [
        [
            "Store Lease - Mall A",
            "Fixed",
            "EGP",
            "Store",
            "Store Name 1",
            "2024-01-01",
            3,
            0,
            "2024-01-01",
            0,
            14,
            "Monthly",
            "No",
            "Increased %",
            5,
            0,
            50000,
            "",
            "",
            "",
            "",
            "",
            0,
            0,
            "none",
            "1,2",
            ""
        ],
        [
            "Revenue Share - Store B",
            "Revenue Share",
            "EGP",
            "Store",
            "Store Name 2",
            "2024-02-01",
            2,
            6,
            "2024-02-01",
            0,
            14,
            "Monthly",
            "No",
            "Increased %",
            3,
            0,
            "",
            100000,
            200000,
            10,
            5,
            "Net",
            0,
            0,
            "none",
            "1",
            ""
        ],
        [
            "ROU Lease - Asset C",
            "ROU",
            "EGP",
            "Other",
            "Asset Name 1",
            "2024-03-01",
            5,
            0,
            "2024-03-01",
            10,
            14,
            "Monthly",
            "Yes",
            "Increased %",
            4,
            0,
            30000,
            "",
            "",
            "",
            "",
            "",
            0,
            0,
            "",
            "1,2",
            "3,4"
        ]
    ]
    
    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_contracts.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            if col_num <= 8:  # Required fields (first 8 columns)
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Create Lessors sheet
    ws_lessors = wb.create_sheet("Lessors", 1)
    lessor_headers = [
        "Contract Name*",
        "Lessor Name*",
        "Share %*"
    ]
    
    for col_num, header in enumerate(lessor_headers, 1):
        cell = ws_lessors.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    ws_lessors.column_dimensions['A'].width = 30
    ws_lessors.column_dimensions['B'].width = 25
    ws_lessors.column_dimensions['C'].width = 15
    
    # Add example lessor data
    lessor_examples = [
        ["Store Lease - Mall A", "Lessor 1", 60],
        ["Store Lease - Mall A", "Lessor 2", 40],
        ["Revenue Share - Store B", "Lessor 1", 100],
        ["ROU Lease - Asset C", "Lessor 1", 50],
        ["ROU Lease - Asset C", "Lessor 2", 50]
    ]
    
    for row_num, row_data in enumerate(lessor_examples, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_lessors.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Create Services sheet (optional)
    ws_services = wb.create_sheet("Services", 2)
    service_headers = [
        "Contract Name*",
        "Service Name*",
        "Amount*",
        "Currency* (EGP/USD)",
        "Yearly Increase %"
    ]
    
    for col_num, header in enumerate(service_headers, 1):
        cell = ws_services.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    ws_services.column_dimensions['A'].width = 30
    ws_services.column_dimensions['B'].width = 25
    ws_services.column_dimensions['C'].width = 20
    ws_services.column_dimensions['D'].width = 20
    ws_services.column_dimensions['E'].width = 20
    
    # Add example service data
    service_examples = [
        ["Store Lease - Mall A", "Maintenance", 5000, "EGP", 3],
        ["Store Lease - Mall A", "Security", 3000, "USD", 0]
    ]
    
    for row_num, row_data in enumerate(service_examples, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_services.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Create Service Lessors sheet (optional)
    ws_service_lessors = wb.create_sheet("Service Lessors", 3)
    service_lessor_headers = [
        "Contract Name*",
        "Service Name*",
        "Lessor Name*",
        "Share %*"
    ]
    
    for col_num, header in enumerate(service_lessor_headers, 1):
        cell = ws_service_lessors.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    ws_service_lessors.column_dimensions['A'].width = 30
    ws_service_lessors.column_dimensions['B'].width = 25
    ws_service_lessors.column_dimensions['C'].width = 25
    ws_service_lessors.column_dimensions['D'].width = 15
    
    # Add example service lessor data
    service_lessor_examples = [
        ["Store Lease - Mall A", "Maintenance", "Lessor 1", 60],
        ["Store Lease - Mall A", "Maintenance", "Lessor 2", 40],
        ["Store Lease - Mall A", "Security", "Lessor 1", 100]
    ]
    
    for row_num, row_data in enumerate(service_lessor_examples, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_service_lessors.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Create Instructions sheet
    ws_instructions = wb.create_sheet("Instructions", 0)
    instructions = [
        ["BULK CONTRACT IMPORT TEMPLATE - INSTRUCTIONS"],
        [""],
        ["1. CONTRACTS SHEET:"],
        ["   - Fill in all required fields (marked with *)"],
        ["   - Contract Name: Unique name for each contract"],
        ["   - Contract Type: Fixed, Revenue Share, or ROU"],
        ["   - Asset/Store Name: Must match existing asset or store name in the system"],
        ["   - Dates: Use YYYY-MM-DD format (e.g., 2024-01-01)"],
        ["   - Tenure: Enter years (0-99) and months (0-11)"],
        ["   - Yearly Increase Type: 'Increased %' or 'Fixed Amount Increased'"],
        ["   - If 'Increased %': Fill Yearly Increase (%) field"],
        ["   - If 'Fixed Amount Increased': Fill Yearly Increase Fixed Amount (EGP) field"],
        ["   - Note: Withholding tax is now handled per lessor via withholding periods (not at contract level)"],
        ["   - For ROU contracts: Discount Rate is required"],
        ["   - For Revenue Share: Revenue fields are required"],
        ["   - For Fixed/ROU: Rent Amount is required"],
        ["   - Advance Payment (Fixed only): optional; same meaning as Create Contract (consumed against rent over time). Use 0 if none."],
        ["   - Revenue Share Payment Advance: optional; prepaid amount applied to payment lines only (chronological). Use 0 if none."],
        ["   - Yearly increase columns map to the same “All periods” increase mode as Create Contract (percent or fixed amount per period)."],
        ["   - Free Months: Available for Fixed, Revenue Share, and ROU (comma-separated period numbers)"],
        ["   - Advance Months: Available for ROU only (comma-separated period numbers)"],
        [""],
        ["2. LESSORS SHEET:"],
        ["   - One row per lessor per contract"],
        ["   - Contract Name must match exactly with Contracts sheet"],
        ["   - Lessor Name must match existing lessor in the system"],
        ["   - Share %: Total must equal 100% for each contract"],
        ["   - Example: Contract with 2 lessors at 60% and 40% = 2 rows"],
        [""],
        ["3. SERVICES SHEET (Optional):"],
        ["   - One row per service per contract"],
        ["   - Contract Name must match exactly with Contracts sheet"],
        ["   - Service Name must match existing service in the system"],
        ["   - Amount: Monthly service amount"],
        ["   - Currency: EGP or USD (must match service's currency in system)"],
        ["   - Yearly Increase %: Optional, default is 0"],
        [""],
        ["4. SERVICE LESSORS SHEET (Optional):"],
        ["   - Assign lessors to services with percentage shares"],
        ["   - One row per lessor per service per contract"],
        ["   - Contract Name and Service Name must match exactly with Contracts and Services sheets"],
        ["   - Lessor Name must match existing lessor in the system"],
        ["   - Share %: Total must equal 100% for each service"],
        ["   - If not provided, service will use contract lessors with equal shares"],
        [""],
        ["5. VALIDATION:"],
        ["   - All required fields must be filled"],
        ["   - Asset/Store and Lessor names must exist in system"],
        ["   - Lessor shares must total 100% per contract"],
        ["   - Service lessor shares must total 100% per service (if provided)"],
        ["   - Contract types must be valid"],
        ["   - Dates must be in correct format"],
        ["   - Currency must match service's currency in system"],
        [""],
        ["6. IMPORT PROCESS:"],
        ["   - Upload the completed Excel file"],
        ["   - System will validate all data"],
        ["   - Contracts will be created in bulk"],
        ["   - Errors will be reported for invalid rows"],
        [""],
        ["NOTE: Delete example rows before adding your data!"]
    ]
    
    for row_num, instruction in enumerate(instructions, 1):
        cell = ws_instructions.cell(row=row_num, column=1)
        cell.value = instruction[0]
        if row_num == 1:
            cell.font = Font(bold=True, size=14)
        elif instruction[0].startswith(("1.", "2.", "3.", "4.", "5.", "6.")):
            cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
        ws_instructions.column_dimensions['A'].width = 100
    
    return wb

def validate_contract_row(row, row_num, lessors_df, assets_df, stores_df):
    """Validate a single contract row"""
    errors = []
    warnings = []
    
    # Required fields
    required_fields = {
        'Contract Name': row.get('Contract Name*', ''),
        'Contract Type': row.get('Contract Type* (Fixed/Revenue Share/ROU)', ''),
        'Currency': row.get('Currency* (EGP/USD)', ''),
        'Asset Category': row.get('Asset Category* (Store/Other)', ''),
        'Asset/Store Name': row.get('Asset/Store Name*', ''),
        'Commencement Date': row.get('Commencement Date* (YYYY-MM-DD)', ''),
        'Tenure Years': row.get('Tenure Years*', ''),
        'Tenure Months': row.get('Tenure Months* (0-11)', ''),
        'Payment Frequency': row.get('Payment Frequency* (Yearly/2 Months/Monthly/Quarter)', '')
    }
    
    for field_name, value in required_fields.items():
        if pd.isna(value) or str(value).strip() == '':
            errors.append(f"Row {row_num}: {field_name} is required")
    
    # Validate contract type
    contract_type = str(required_fields['Contract Type']).strip()
    if contract_type not in ['Fixed', 'Revenue Share', 'ROU']:
        errors.append(f"Row {row_num}: Invalid Contract Type '{contract_type}'. Must be Fixed, Revenue Share, or ROU")
    
    # Validate currency
    currency = str(required_fields['Currency']).strip()
    if currency not in ['EGP', 'USD']:
        errors.append(f"Row {row_num}: Invalid Currency '{currency}'. Must be EGP or USD")
    
    # Validate asset category
    asset_category = str(required_fields['Asset Category']).strip()
    if asset_category not in ['Store', 'Other']:
        errors.append(f"Row {row_num}: Invalid Asset Category '{asset_category}'. Must be Store or Other")
    
    # Validate asset/store exists (case-insensitive)
    asset_store_name = str(required_fields['Asset/Store Name']).strip()
    if asset_category == 'Store':
        if stores_df.empty:
            errors.append(f"Row {row_num}: No stores found in system")
        else:
            store_match = stores_df[stores_df['name'].str.strip().str.lower() == asset_store_name.lower()]
            if store_match.empty:
                errors.append(f"Row {row_num}: Store '{asset_store_name}' not found in system")
    else:
        if assets_df.empty:
            errors.append(f"Row {row_num}: No assets found in system")
        else:
            asset_match = assets_df[assets_df['name'].str.strip().str.lower() == asset_store_name.lower()]
            if asset_match.empty:
                errors.append(f"Row {row_num}: Asset '{asset_store_name}' not found in system")
    
    # Validate dates
    try:
        commencement_date = pd.to_datetime(required_fields['Commencement Date'])
    except:
        errors.append(f"Row {row_num}: Invalid Commencement Date format. Use YYYY-MM-DD")
    
    # Validate tenure
    try:
        tenure_years = int(float(required_fields['Tenure Years']))
        tenure_months_only = int(float(required_fields['Tenure Months']))
        if tenure_years < 0 or tenure_months_only < 0 or tenure_months_only > 11:
            errors.append(f"Row {row_num}: Invalid tenure. Years >= 0, Months 0-11")
    except:
        errors.append(f"Row {row_num}: Invalid tenure values")
    
    # Validate payment frequency
    payment_freq = str(required_fields['Payment Frequency']).strip()
    if payment_freq not in ['Yearly', '2 Months', 'Monthly', 'Quarter']:
        errors.append(f"Row {row_num}: Invalid Payment Frequency '{payment_freq}'")
    
    # Contract type specific validations
    if contract_type == 'ROU':
        discount_rate = row.get('Discount Rate (%) (ROU only)', 0)
        try:
            discount_rate = float(discount_rate) if not pd.isna(discount_rate) else 0
            if discount_rate <= 0:
                errors.append(f"Row {row_num}: Discount Rate is required and must be > 0 for ROU contracts")
        except:
            errors.append(f"Row {row_num}: Invalid Discount Rate value")
    
    if contract_type in ['Fixed', 'ROU']:
        rent_amount = row.get('Rent Amount (EGP/month) (Fixed/ROU)', 0)
        try:
            rent_amount = float(rent_amount) if not pd.isna(rent_amount) else 0
            if rent_amount <= 0:
                errors.append(f"Row {row_num}: Rent Amount is required and must be > 0 for {contract_type} contracts")
        except:
            errors.append(f"Row {row_num}: Invalid Rent Amount value")
    
    if contract_type == 'Revenue Share':
        if asset_category != 'Store':
            errors.append(f"Row {row_num}: Revenue Share contracts are only available for Store assets")
        rev_min = row.get('Revenue Min (EGP) (Revenue Share)', 0)
        rev_share_pct = row.get('Revenue Share % (Revenue Share)', 0)
        try:
            rev_min = float(rev_min) if not pd.isna(rev_min) else 0
            rev_share_pct = float(rev_share_pct) if not pd.isna(rev_share_pct) else 0
            if rev_min <= 0 or rev_share_pct <= 0:
                errors.append(f"Row {row_num}: Revenue Min and Revenue Share % are required for Revenue Share contracts")
        except:
            errors.append(f"Row {row_num}: Invalid Revenue Share values")
    
    for _lbl, _raw in (
        (_HDR_ADVANCE_PAYMENT_FIXED, row.get(_HDR_ADVANCE_PAYMENT_FIXED, 0)),
        (_HDR_RS_PAYMENT_ADVANCE, row.get(_HDR_RS_PAYMENT_ADVANCE, 0)),
    ):
        try:
            _v = float(_raw) if not pd.isna(_raw) else 0.0
            if _v < 0:
                errors.append(f"Row {row_num}: {_lbl} cannot be negative")
        except Exception:
            errors.append(f"Row {row_num}: Invalid numeric value for {_lbl}")
    
    return errors, warnings

def validate_lessor_row(row, row_num, lessors_df, contract_names):
    """Validate a lessor row"""
    errors = []
    
    contract_name = str(row.get('Contract Name*', '')).strip()
    lessor_name = str(row.get('Lessor Name*', '')).strip()
    share_pct = row.get('Share %*', '')
    
    if pd.isna(contract_name) or contract_name == '':
        errors.append(f"Lessors Row {row_num}: Contract Name is required")
    elif contract_name not in contract_names:
        errors.append(f"Lessors Row {row_num}: Contract Name '{contract_name}' not found in Contracts sheet")
    
    if pd.isna(lessor_name) or lessor_name == '':
        errors.append(f"Lessors Row {row_num}: Lessor Name is required")
    elif lessors_df.empty:
        errors.append(f"Lessors Row {row_num}: No lessors found in system")
    else:
        lessor_match = lessors_df[lessors_df['name'].str.strip().str.lower() == lessor_name.lower()]
        if lessor_match.empty:
            errors.append(f"Lessors Row {row_num}: Lessor '{lessor_name}' not found in system")
    
    try:
        share_pct = float(share_pct) if not pd.isna(share_pct) else 0
        if share_pct <= 0 or share_pct > 100:
            errors.append(f"Lessors Row {row_num}: Share % must be between 0 and 100")
    except:
        errors.append(f"Lessors Row {row_num}: Invalid Share % value")
    
    return errors

def validate_service_row(row, row_num, services_df, contract_names):
    """Validate a service row"""
    errors = []
    
    contract_name = str(row.get('Contract Name*', '')).strip()
    service_name = str(row.get('Service Name*', '')).strip()
    amount = row.get('Amount*', '')
    currency = str(row.get('Currency* (EGP/USD)', '')).strip()
    
    if pd.isna(contract_name) or contract_name == '':
        errors.append(f"Services Row {row_num}: Contract Name is required")
    elif contract_name not in contract_names:
        errors.append(f"Services Row {row_num}: Contract Name '{contract_name}' not found in Contracts sheet")
    
    if pd.isna(service_name) or service_name == '':
        errors.append(f"Services Row {row_num}: Service Name is required")
    elif services_df.empty:
        errors.append(f"Services Row {row_num}: No services found in system")
    else:
        service_match = services_df[services_df['name'].str.strip().str.lower() == service_name.lower()]
        if service_match.empty:
            errors.append(f"Services Row {row_num}: Service '{service_name}' not found in system")
        else:
            # Validate currency matches service's currency
            service_currency = service_match.iloc[0].get('currency', 'EGP')
            if currency.upper() != service_currency.upper():
                errors.append(f"Services Row {row_num}: Currency '{currency}' does not match service's currency '{service_currency}'")
    
    if pd.isna(currency) or currency == '':
        errors.append(f"Services Row {row_num}: Currency is required")
    elif currency.upper() not in ['EGP', 'USD']:
        errors.append(f"Services Row {row_num}: Invalid Currency '{currency}'. Must be EGP or USD")
    
    try:
        amount = float(amount) if not pd.isna(amount) else 0
        if amount <= 0:
            errors.append(f"Services Row {row_num}: Amount must be > 0")
    except:
        errors.append(f"Services Row {row_num}: Invalid Amount value")
    
    return errors

def validate_service_lessor_row(row, row_num, lessors_df, services_df, contract_names, service_names):
    """Validate a service lessor row"""
    errors = []
    
    contract_name = str(row.get('Contract Name*', '')).strip()
    service_name = str(row.get('Service Name*', '')).strip()
    lessor_name = str(row.get('Lessor Name*', '')).strip()
    share_pct = row.get('Share %*', '')
    
    if pd.isna(contract_name) or contract_name == '':
        errors.append(f"Service Lessors Row {row_num}: Contract Name is required")
    elif contract_name not in contract_names:
        errors.append(f"Service Lessors Row {row_num}: Contract Name '{contract_name}' not found in Contracts sheet")
    
    if pd.isna(service_name) or service_name == '':
        errors.append(f"Service Lessors Row {row_num}: Service Name is required")
    elif service_name not in service_names:
        errors.append(f"Service Lessors Row {row_num}: Service Name '{service_name}' not found in Services sheet")
    
    if pd.isna(lessor_name) or lessor_name == '':
        errors.append(f"Service Lessors Row {row_num}: Lessor Name is required")
    elif lessors_df.empty:
        errors.append(f"Service Lessors Row {row_num}: No lessors found in system")
    else:
        lessor_match = lessors_df[lessors_df['name'].str.strip().str.lower() == lessor_name.lower()]
        if lessor_match.empty:
            errors.append(f"Service Lessors Row {row_num}: Lessor '{lessor_name}' not found in system")
    
    try:
        share_pct = float(share_pct) if not pd.isna(share_pct) else 0
        if share_pct <= 0 or share_pct > 100:
            errors.append(f"Service Lessors Row {row_num}: Share % must be between 0 and 100")
    except:
        errors.append(f"Service Lessors Row {row_num}: Invalid Share % value")
    
    return errors

def process_bulk_import(uploaded_file):
    """Process bulk import Excel file"""
    try:
        # Load all dataframes
        load_all()
        lessors_df = st.session_state.lessors_df.copy()
        assets_df = st.session_state.assets_df.copy()
        stores_df = st.session_state.stores_df.copy()
        services_df = st.session_state.services_df.copy()
        contracts_df = st.session_state.contracts_df.copy()
        
        # Read Excel sheets
        contracts_sheet = pd.read_excel(uploaded_file, sheet_name='Contracts')
        lessors_sheet = pd.read_excel(uploaded_file, sheet_name='Lessors')
        
        # Try to read services sheet (optional)
        try:
            services_sheet = pd.read_excel(uploaded_file, sheet_name='Services')
        except:
            services_sheet = pd.DataFrame()
        
        all_errors = []
        all_warnings = []
        
        # Validate contracts
        contract_names = []
        for idx, row in contracts_sheet.iterrows():
            row_num = idx + 2  # Excel row number (1-based + header)
            if _excel_row_effectively_empty(row.values):
                continue
            errors, warnings = validate_contract_row(row, row_num, lessors_df, assets_df, stores_df)
            all_errors.extend(errors)
            all_warnings.extend(warnings)
            
            contract_name = str(row.get('Contract Name*', '')).strip()
            if contract_name and not pd.isna(contract_name):
                contract_names.append(contract_name)
        
        # Validate lessors
        lessor_groups = {}
        for idx, row in lessors_sheet.iterrows():
            row_num = idx + 2
            errors = validate_lessor_row(row, row_num, lessors_df, contract_names)
            all_errors.extend(errors)
            
            if not errors:
                contract_name = str(row.get('Contract Name*', '')).strip()
                lessor_name = str(row.get('Lessor Name*', '')).strip()
                share_pct = float(row.get('Share %*', 0))
                
                if contract_name not in lessor_groups:
                    lessor_groups[contract_name] = []
                lessor_groups[contract_name].append({
                    'name': lessor_name,
                    'share': share_pct
                })
        
        # Validate lessor shares total 100%
        for contract_name, lessors in lessor_groups.items():
            total_share = sum(l['share'] for l in lessors)
            if abs(total_share - 100.0) > 0.01:
                all_errors.append(f"Contract '{contract_name}': Lessor shares total {total_share:.2f}%, must equal 100%")
        
        # Validate services
        service_groups = {}
        service_names_by_contract = {}  # Track service names per contract for service lessor validation
        if not services_sheet.empty:
            for idx, row in services_sheet.iterrows():
                row_num = idx + 2
                errors = validate_service_row(row, row_num, services_df, contract_names)
                all_errors.extend(errors)
                
                if not errors:
                    contract_name = str(row.get('Contract Name*', '')).strip()
                    service_name = str(row.get('Service Name*', '')).strip()
                    amount = float(row.get('Amount*', 0))
                    yearly_increase = float(row.get('Yearly Increase %', 0)) if not pd.isna(row.get('Yearly Increase %', 0)) else 0
                    
                    if contract_name not in service_groups:
                        service_groups[contract_name] = []
                        service_names_by_contract[contract_name] = []
                    service_groups[contract_name].append({
                        'name': service_name,
                        'amount': amount,
                        'yearly_increase_pct': yearly_increase
                    })
                    service_names_by_contract[contract_name].append(service_name)
        
        # Validate service lessors (optional)
        service_lessor_groups = {}  # {contract_name: {service_name: [lessors]}}
        try:
            service_lessors_sheet = pd.read_excel(uploaded_file, sheet_name='Service Lessors')
        except:
            service_lessors_sheet = pd.DataFrame()
        
        if not service_lessors_sheet.empty:
            for idx, row in service_lessors_sheet.iterrows():
                row_num = idx + 2
                contract_name = str(row.get('Contract Name*', '')).strip()
                service_name = str(row.get('Service Name*', '')).strip()
                service_names = service_names_by_contract.get(contract_name, [])
                errors = validate_service_lessor_row(row, row_num, lessors_df, services_df, contract_names, service_names)
                all_errors.extend(errors)
                
                if not errors:
                    lessor_name = str(row.get('Lessor Name*', '')).strip()
                    share_pct = float(row.get('Share %*', 0))
                    
                    if contract_name not in service_lessor_groups:
                        service_lessor_groups[contract_name] = {}
                    if service_name not in service_lessor_groups[contract_name]:
                        service_lessor_groups[contract_name][service_name] = []
                    service_lessor_groups[contract_name][service_name].append({
                        'name': lessor_name,
                        'share': share_pct
                    })
        
        # Validate service lessor shares total 100% per service
        for contract_name, services_dict in service_lessor_groups.items():
            for service_name, lessors in services_dict.items():
                total_share = sum(l['share'] for l in lessors)
                if abs(total_share - 100.0) > 0.01:
                    all_errors.append(f"Contract '{contract_name}', Service '{service_name}': Service lessor shares total {total_share:.2f}%, must equal 100%")
        
        # If there are errors, return them
        if all_errors:
            return False, all_errors, all_warnings, 0
        
        # Process valid contracts
        success_count = 0
        failed_contracts = []
        
        for idx, row in contracts_sheet.iterrows():
            try:
                if _excel_row_effectively_empty(row.values):
                    continue
                contract_name = str(row.get('Contract Name*', '')).strip()
                if not contract_name:
                    continue
                contract_type = str(row.get('Contract Type* (Fixed/Revenue Share/ROU)', '')).strip()
                currency = str(row.get('Currency* (EGP/USD)', '')).strip()
                asset_category = str(row.get('Asset Category* (Store/Other)', '')).strip()
                asset_store_name = str(row.get('Asset/Store Name*', '')).strip()
                
                # Get asset/store ID (case-insensitive)
                if asset_category == 'Store':
                    asset_store_row = stores_df[stores_df['name'].str.strip().str.lower() == asset_store_name.lower()]
                else:
                    asset_store_row = assets_df[assets_df['name'].str.strip().str.lower() == asset_store_name.lower()]
                
                if asset_store_row.empty:
                    failed_contracts.append(f"{contract_name}: Asset/Store not found")
                    continue
                
                asset_store_id = str(asset_store_row.iloc[0]['id'])
                
                # Dates
                commencement_date = pd.to_datetime(row.get('Commencement Date* (YYYY-MM-DD)')).date()
                tenure_years = int(float(row.get('Tenure Years*', 0)))
                tenure_months_only = int(float(row.get('Tenure Months* (0-11)', 0)))
                tenure_months = tenure_years * 12 + tenure_months_only
                end_date_iso = calc_end_date_iso(commencement_date, tenure_months)
                
                first_payment_date_str = row.get('First Payment Date (YYYY-MM-DD)', '')
                if pd.isna(first_payment_date_str) or str(first_payment_date_str).strip() == '':
                    first_payment_date = commencement_date
                else:
                    first_payment_date = pd.to_datetime(first_payment_date_str).date()
                
                # Financial fields
                discount_rate = float(row.get('Discount Rate (%) (ROU only)', 0)) if not pd.isna(row.get('Discount Rate (%) (ROU only)', 0)) else 0
                tax_per = float(row.get('Tax (%)', 0)) if not pd.isna(row.get('Tax (%)', 0)) else 0
                # Note: Holding tax is no longer stored at contract level - it's handled per lessor via withholding periods
                payment_frequency = str(row.get('Payment Frequency* (Yearly/2 Months/Monthly/Quarter)', '')).strip()
                is_tax_added_raw = str(row.get('Is Tax Added (ROU: Yes/No)', 'No')).strip().lower()
                is_tax_added = 1 if is_tax_added_raw in ['yes', 'y', '1', 'true'] else 0
                
                # Yearly increase type and values
                yearly_increase_type = str(row.get('Yearly Increase Type (Increased %/Fixed Amount Increased)', 'Increased %')).strip()
                if yearly_increase_type not in ['Increased %', 'Fixed Amount Increased']:
                    yearly_increase_type = 'Increased %'
                
                if yearly_increase_type == 'Increased %':
                    yearly_increase = float(row.get('Yearly Increase (%)', 0)) if not pd.isna(row.get('Yearly Increase (%)', 0)) else 0
                    yearly_increase_fixed_amount = 0
                else:
                    yearly_increase = 0
                    _yr_fix_key = row.get(
                        "Yearly Increase Fixed Amount (per period, same currency as contract)",
                        row.get("Yearly Increase Fixed Amount (EGP)", 0),
                    )
                    yearly_increase_fixed_amount = float(_yr_fix_key) if not pd.isna(_yr_fix_key) else 0
                
                if yearly_increase_type == 'Increased %':
                    increase_by_period_all_pct = str(yearly_increase)
                    increase_by_period_map = json.dumps({"all_value_type": "percent"})
                else:
                    increase_by_period_all_pct = str(yearly_increase_fixed_amount)
                    increase_by_period_map = json.dumps({"all_value_type": "amount"})
                
                rent_amount = float(row.get('Rent Amount (EGP/month) (Fixed/ROU)', 0)) if not pd.isna(row.get('Rent Amount (EGP/month) (Fixed/ROU)', 0)) else 0
                
                # Revenue Share fields
                rev_min = float(row.get('Revenue Min (EGP) (Revenue Share)', 0)) if not pd.isna(row.get('Revenue Min (EGP) (Revenue Share)', 0)) else 0
                rev_max = float(row.get('Revenue Max (EGP) (Revenue Share)', 0)) if not pd.isna(row.get('Revenue Max (EGP) (Revenue Share)', 0)) else 0
                rev_share_pct = float(row.get('Revenue Share % (Revenue Share)', 0)) if not pd.isna(row.get('Revenue Share % (Revenue Share)', 0)) else 0
                rev_share_after_max_pc = float(row.get('Share % After Max (Revenue Share)', 0)) if not pd.isna(row.get('Share % After Max (Revenue Share)', 0)) else 0
                sales_type = str(row.get('Sales Type (Revenue Share: Net/Total without discount)', '')).strip() if contract_type == 'Revenue Share' else ''
                
                # Free months (available for Fixed, Revenue Share, and ROU)
                free_months = str(row.get('Free Months (comma-separated, Fixed/Revenue Share/ROU: e.g., 1,2,3)', '')).strip()
                if contract_type not in ['Fixed', 'Revenue Share', 'ROU']:
                    free_months = ''
                
                # Advance months (ROU only)
                _adv_hdr = "Advance Months (ROU / Revenue Share “periods” advance mode, e.g. 1,2,3)"
                advance_months = (
                    str(row.get(_adv_hdr, row.get('Advance Months (comma-separated, ROU: e.g., 3,4)', ''))).strip()
                    if contract_type in ('ROU', 'Revenue Share')
                    else ''
                )
                advance_months_count = ""
                if contract_type in ('ROU', 'Revenue Share') and advance_months:
                    advance_months_count = str(
                        len([x for x in advance_months.split(",") if x.strip().isdigit()])
                    )
                adv_pay = (
                    float(row.get(_HDR_ADVANCE_PAYMENT_FIXED, 0))
                    if not pd.isna(row.get(_HDR_ADVANCE_PAYMENT_FIXED, 0))
                    else 0.0
                )
                rs_adv = (
                    float(row.get(_HDR_RS_PAYMENT_ADVANCE, 0))
                    if not pd.isna(row.get(_HDR_RS_PAYMENT_ADVANCE, 0))
                    else 0.0
                )
                rs_adv_mode_raw = str(row.get(_HDR_RS_ADV_MODE, "none") or "none").strip().lower()
                if rs_adv_mode_raw in ("", "legacy"):
                    rs_adv_mode_raw = "none"
                if rs_adv_mode_raw not in ("none", "chronological", "periods", "spread_proportional"):
                    rs_adv_mode_raw = "none"
                rent_per_year_json = ""  # Will be calculated if needed
                
                # Get lessors for this contract
                contract_lessors = lessor_groups.get(contract_name, [])
                if not contract_lessors:
                    failed_contracts.append(f"{contract_name}: No lessors assigned")
                    continue
                
                # Build lessors JSON and get lessor IDs (case-insensitive)
                lessors_json_list = []
                for lessor_info in contract_lessors:
                    lessor_row = lessors_df[lessors_df['name'].str.strip().str.lower() == lessor_info['name'].lower()]
                    if not lessor_row.empty:
                        lessor_id = str(lessor_row.iloc[0]['id'])
                        lessor_actual_name = lessor_row.iloc[0]['name']  # Use actual name from DB
                        lessors_json_list.append({
                            'id': lessor_id,
                            'name': lessor_actual_name,
                            'share': lessor_info['share'],
                            'supplier_code': lessor_row.iloc[0].get('supplier_code', '')
                        })
                
                lessors_json = json.dumps(lessors_json_list, ensure_ascii=False)
                
                # Generate contract ID
                nid = next_int_id(contracts_df, 1001)
                
                # Build contract data
                contract_data = {
                    "id": str(nid),
                    "contract_name": contract_name,
                    "contract_type": contract_type,
                    "currency": currency,
                    "asset_category": asset_category,
                    "asset_or_store_id": asset_store_id,
                    "asset_or_store_name": asset_store_name,
                    "commencement_date": str(commencement_date),
                    "tenure_months": str(tenure_months),
                    "end_date": end_date_iso,
                    "lessors_json": lessors_json,
                    "discount_rate": str(discount_rate),
                    "tax": str(tax_per),
                    "is_tax_added": str(is_tax_added if contract_type == "ROU" else 0),
                    # Note: holding_tax removed - withholding tax is now handled per lessor via withholding periods
                    "payment_frequency": payment_frequency,
                    "yearly_increase": str(yearly_increase),
                    "yearly_increase_type": yearly_increase_type,
                    "yearly_increase_fixed_amount": str(yearly_increase_fixed_amount),
                    "rent_amount": str(rent_amount),
                    "rev_min": str(rev_min),
                    "rev_max": str(rev_max),
                    "rev_share_pct": str(rev_share_pct),
                    "rev_share_after_max_pc": str(rev_share_after_max_pc),
                    "sales_type": sales_type,
                    "rent_per_year": rent_per_year_json,
                    "first_payment_date": str(first_payment_date) if first_payment_date else "",
                    "free_months": free_months,
                    "advance_months": advance_months,
                    "advance_months_count": advance_months_count,
                    "increase_by_period_mode": "all",
                    "increase_by_period_all_pct": increase_by_period_all_pct,
                    "increase_by_period_map": increase_by_period_map,
                    "advance_payment": str(adv_pay) if contract_type == "Fixed" else "",
                    "rev_share_payment_advance": str(rs_adv) if contract_type == "Revenue Share" else "",
                    "rev_share_advance_mode": rs_adv_mode_raw if contract_type == "Revenue Share" else "",
                    "created_at": time.strftime("%Y-%m-%d %H:%M:%S")
                }
                
                # Insert contract
                if insert_contract(contract_data):
                    # Log action for bulk import
                    current_user = get_current_user()
                    log_action(
                        user_id=current_user['id'] if current_user else None,
                        user_name=current_user['name'] if current_user else 'System',
                        action_type='create',
                        entity_type='contract',
                        entity_id=str(nid),
                        entity_name=contract_name,
                        action_details=f"Bulk imported {contract_type} contract: {contract_name}",
                        ip_address=get_user_ip()
                    )
                    # Insert contract-lessor relationships (case-insensitive)
                    for lessor_info in contract_lessors:
                        lessor_row = lessors_df[lessors_df['name'].str.strip().str.lower() == lessor_info['name'].lower()]
                        if not lessor_row.empty:
                            lessor_id = str(lessor_row.iloc[0]['id'])
                            insert_contract_lessor(str(nid), lessor_id, str(lessor_info['share']))
                    
                    # Insert contract-service relationships (case-insensitive)
                    contract_services = service_groups.get(contract_name, [])
                    for service_info in contract_services:
                        service_row = services_df[services_df['name'].str.strip().str.lower() == service_info['name'].lower()]
                        if not service_row.empty:
                            service_id = str(service_row.iloc[0]['id'])
                            insert_contract_service(
                                str(nid),
                                service_id,
                                str(service_info['amount']),
                                str(service_info.get('yearly_increase_pct', 0))
                            )
                            
                            # Insert service-lessor relationships if provided
                            service_lessors = service_lessor_groups.get(contract_name, {}).get(service_info['name'], [])
                            if service_lessors:
                                # Validate shares total 100%
                                total_share = sum(l['share'] for l in service_lessors)
                                if abs(total_share - 100.0) <= 0.01:
                                    for lessor_info in service_lessors:
                                        lessor_row = lessors_df[lessors_df['name'].str.strip().str.lower() == lessor_info['name'].lower()]
                                        if not lessor_row.empty:
                                            lessor_id = str(lessor_row.iloc[0]['id'])
                                            insert_contract_service_lessor(
                                                str(nid),
                                                service_id,
                                                lessor_id,
                                                str(lessor_info['share'])
                                            )
                    
                    success_count += 1
                else:
                    failed_contracts.append(f"{contract_name}: Database insert failed")
            
            except Exception as e:
                contract_name = str(row.get('Contract Name*', 'Unknown')).strip()
                failed_contracts.append(f"{contract_name}: {str(e)}")
        
        # Reload data
        load_all()
        
        return True, failed_contracts, all_warnings, success_count
        
    except Exception as e:
        return False, [f"Error processing file: {str(e)}"], [], 0

def _render_bulk_contract_import():
    """Contracts + contract-level lessors/services (existing flow)."""
    st.subheader("Contracts import")
    
    st.markdown("""
    ### Instructions
    1. Download the Excel template below
    2. Fill in contract data in the **Contracts** sheet
    3. Add lessors in the **Lessors** sheet (one row per lessor per contract)
    4. Optionally add services in the **Services** sheet (with currency)
    5. Optionally assign lessors to services in the **Service Lessors** sheet
    6. Upload the completed file to import contracts in bulk
    
    **Important Notes:**
    - All required fields (marked with *) must be filled
    - Asset/Store names and **Lessor names** must match **existing** master records in the system
    - To create lessors, assets, or services in bulk first, use **Bulk Import** → **Lessors**, **Assets**, or **Services**
    - Lessor shares must total exactly 100% for each contract
    - Service lessor shares must total exactly 100% for each service (if provided)
    - Service currency must match the service's currency in the system
    - Free months are available for Fixed, Revenue Share, and ROU contracts
    - Yearly Increase Type: Choose 'Increased %' or 'Fixed Amount Increased' (same **All periods** logic as Create Contract)
    - **Advance Payment (Fixed only)** and **Revenue Share Payment Advance** are optional (use 0 if not applicable)
    - Delete example rows before adding your data
    """)
    
    # Generate and download template
    st.markdown("### Step 1: Download Template")
    
    # Generate template
    wb = generate_bulk_import_template()
    
    # Save to bytes
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    st.download_button(
        label="📥 Download Excel Template",
        data=output.getvalue(),
        file_name=f"Contract_Bulk_Import_Template_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_template_file"
    )
    
    st.markdown("---")
    st.markdown("### Step 2: Upload Completed File")
    
    uploaded_file = st.file_uploader(
        "Choose Excel file",
        type=['xlsx', 'xls'],
        key="bulk_import_file"
    )
    
    if uploaded_file is not None:
        if st.button("Import Contracts", key="import_contracts", type="primary"):
            from core.permissions import require_permission
            require_permission('bulk_import.import')
            with st.spinner("Processing import..."):
                success, errors, warnings, success_count = process_bulk_import(uploaded_file)
            
            if success and success_count > 0:
                # Log bulk import completion
                current_user = get_current_user()
                log_action(
                    user_id=current_user['id'] if current_user else None,
                    user_name=current_user['name'] if current_user else 'System',
                    action_type='bulk_import',
                    entity_type='contract',
                    entity_id=None,
                    entity_name=None,
                    action_details=f"Bulk import completed: {success_count} contract(s) imported",
                    ip_address=get_user_ip()
                )
                st.success(f"✅ Successfully imported {success_count} contract(s)!")
                
                if warnings:
                    st.warning("⚠️ Warnings:")
                    for warning in warnings:
                        st.write(f"- {warning}")
                
                if errors:
                    st.error("❌ Failed contracts:")
                    for error in errors:
                        st.write(f"- {error}")
            elif success and success_count == 0:
                st.warning("⚠️ No contracts were imported. Check errors below.")
                if errors:
                    st.error("❌ Errors:")
                    for error in errors:
                        st.write(f"- {error}")
            else:
                st.error("❌ Import failed. Please fix the following errors:")
                for error in errors:
                    st.write(f"- {error}")
                
                if warnings:
                    st.warning("⚠️ Warnings:")
                    for warning in warnings:
                        st.write(f"- {warning}")


def _render_bulk_master_import(
    *,
    subheader: str,
    instructions_md: str,
    template_wb,
    download_file_name: str,
    download_key: str,
    uploader_key: str,
    import_button_key: str,
    import_button_label: str,
    process_fn,
    log_entity_type: str,
    log_entity_label: str,
):
    """Download template + upload + import for master-data sheets (lessors / assets / services)."""
    st.subheader(subheader)
    st.markdown(instructions_md)
    st.markdown("### Step 1: Download template")
    from io import BytesIO

    output = BytesIO()
    template_wb.save(output)
    output.seek(0)
    st.download_button(
        label="Download Excel template",
        data=output.getvalue(),
        file_name=download_file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=download_key,
    )
    st.markdown("---")
    st.markdown("### Step 2: Upload completed file")
    uploaded_file = st.file_uploader(
        "Choose Excel file",
        type=["xlsx", "xls"],
        key=uploader_key,
    )
    if uploaded_file is not None:
        if st.button(import_button_label, key=import_button_key, type="primary"):
            from core.permissions import require_permission

            require_permission("bulk_import.import")
            with st.spinner("Processing import..."):
                success, errors, warnings, success_count = process_fn(uploaded_file)

            if success and success_count > 0:
                current_user = get_current_user()
                log_action(
                    user_id=current_user["id"] if current_user else None,
                    user_name=current_user["name"] if current_user else "System",
                    action_type="bulk_import",
                    entity_type=log_entity_type,
                    entity_id=None,
                    entity_name=None,
                    action_details=f"Bulk import completed: {success_count} {log_entity_label}",
                    ip_address=get_user_ip(),
                )
                st.success(f"Successfully imported {success_count} {log_entity_label}!")
                if warnings:
                    st.warning("Warnings:")
                    for warning in warnings:
                        st.write(f"- {warning}")
                if errors:
                    st.error("Rows that did not import:")
                    for error in errors:
                        st.write(f"- {error}")
            elif success and success_count == 0:
                st.warning("No rows were imported. Check errors below.")
                if errors:
                    st.error("Errors:")
                    for error in errors:
                        st.write(f"- {error}")
            else:
                st.error("Import failed. Please fix the following errors:")
                for error in errors:
                    st.write(f"- {error}")
                if warnings:
                    st.warning("Warnings:")
                    for warning in warnings:
                        st.write(f"- {warning}")


def render_bulk_import_tab():
    """Bulk Import hub: contracts plus master lessors, assets, and services."""
    tab_contracts, tab_lessors, tab_assets, tab_services = st.tabs(
        ["Contracts", "Lessors", "Assets", "Services"]
    )
    with tab_contracts:
        _render_bulk_contract_import()
    with tab_lessors:
        _render_bulk_master_import(
            subheader="Lessors import",
            instructions_md="""
### Instructions
1. Download the template and use the **Lessors** sheet only.
2. **Lessor Name*** is required and must be unique (not already in the system).
3. Optional columns: Description, Tax ID, Supplier Code, IBAN.
4. Delete the example rows before importing.
5. Import **lessors** before contract bulk import if those names are not yet in the system.
""",
            template_wb=generate_master_lessors_template(),
            download_file_name=f"Lessor_Bulk_Import_Template_{datetime.now().strftime('%Y%m%d')}.xlsx",
            download_key="download_master_lessors_template",
            uploader_key="bulk_import_master_lessors_file",
            import_button_key="import_master_lessors",
            import_button_label="Import lessors",
            process_fn=process_bulk_import_master_lessors,
            log_entity_type="lessor",
            log_entity_label="lessor(s)",
        )
    with tab_assets:
        _render_bulk_master_import(
            subheader="Assets import",
            instructions_md="""
### Instructions
1. Download the template and use the **Assets** sheet only.
2. **Asset Name*** is required and must be unique (not already in the system).
3. **Cost Center** is optional.
4. Delete the example rows before importing.
5. Import **assets** before contract bulk import if those names are not yet in the system.
""",
            template_wb=generate_master_assets_template(),
            download_file_name=f"Asset_Bulk_Import_Template_{datetime.now().strftime('%Y%m%d')}.xlsx",
            download_key="download_master_assets_template",
            uploader_key="bulk_import_master_assets_file",
            import_button_key="import_master_assets",
            import_button_label="Import assets",
            process_fn=process_bulk_import_master_assets,
            log_entity_type="asset",
            log_entity_label="asset(s)",
        )
    with tab_services:
        _render_bulk_master_import(
            subheader="Services import",
            instructions_md="""
### Instructions
1. Download the template and use the **Services** sheet only.
2. **Service Name*** is required and must be unique (not already in the system).
3. **Currency*** must be **EGP** or **USD** (must match how the service will be used on contracts).
4. Description is optional.
5. Delete the example rows before importing.
6. Import **services** before contract bulk import if those names are not yet in the system.
""",
            template_wb=generate_master_services_template(),
            download_file_name=f"Service_Bulk_Import_Template_{datetime.now().strftime('%Y%m%d')}.xlsx",
            download_key="download_master_services_template",
            uploader_key="bulk_import_master_services_file",
            import_button_key="import_master_services",
            import_button_label="Import services",
            process_fn=process_bulk_import_master_services,
            log_entity_type="service",
            log_entity_label="service(s)",
        )
